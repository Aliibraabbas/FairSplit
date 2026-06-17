from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from balances.services import calculate_balances, calculate_settlements


def export_group_to_excel(group):
    """Exporte toutes les données d'un groupe en Excel avec plusieurs feuilles"""
    wb = Workbook()
    currency_symbol = group.get_currency_symbol()
    
    # Styles
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    title_font = Font(bold=True, size=14)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # === FEUILLE 1 : RÉSUMÉ ===
    ws_summary = wb.active
    ws_summary.title = "Resume"
    
    ws_summary['A1'] = "FAIRSPLIT - RÉSUMÉ DU GROUPE"
    ws_summary['A1'].font = Font(bold=True, size=16, color="4F46E5")
    ws_summary['A3'] = "Nom du groupe:"
    ws_summary['B3'] = group.name
    ws_summary['A4'] = "Description:"
    ws_summary['B4'] = group.description or "Aucune"
    ws_summary['A5'] = "Devise:"
    ws_summary['B5'] = f"{group.get_currency_display()}"
    ws_summary['A6'] = "Créé le:"
    ws_summary['B6'] = group.created_at.strftime('%d/%m/%Y')
    ws_summary['A7'] = "Créé par:"
    ws_summary['B7'] = group.created_by.username
    
    ws_summary['A9'] = "STATISTIQUES"
    ws_summary['A9'].font = title_font
    ws_summary['A10'] = "Nombre de membres:"
    ws_summary['B10'] = group.members.count()
    ws_summary['A11'] = "Nombre d'invités:"
    ws_summary['B11'] = group.guest_members.count()
    ws_summary['A12'] = "Nombre de dépenses:"
    total_expenses = group.expenses.count()
    ws_summary['B12'] = total_expenses
    ws_summary['A13'] = "Total des dépenses:"
    total_amount = sum(float(e.amount) for e in group.expenses.all())
    ws_summary['B13'] = f"{total_amount:.2f} {currency_symbol}"
    
    # Largeur des colonnes
    ws_summary.column_dimensions['A'].width = 25
    ws_summary.column_dimensions['B'].width = 40
    
    # === FEUILLE 2 : MEMBRES ===
    ws_members = wb.create_sheet("Membres")
    ws_members['A1'] = "MEMBRES INSCRITS"
    ws_members['A1'].font = title_font
    
    headers = ['ID', 'Nom d\'utilisateur', 'Email']
    for col, header in enumerate(headers, 1):
        cell = ws_members.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    row = 4
    for member in group.members.all():
        ws_members.cell(row=row, column=1, value=member.id).border = border
        ws_members.cell(row=row, column=2, value=member.username).border = border
        ws_members.cell(row=row, column=3, value=member.email).border = border
        row += 1
    
    ws_members.column_dimensions['A'].width = 10
    ws_members.column_dimensions['B'].width = 25
    ws_members.column_dimensions['C'].width = 35
    
    # === FEUILLE 3 : INVITÉS ===
    ws_guests = wb.create_sheet("Invites")
    ws_guests['A1'] = "INVITÉS"
    ws_guests['A1'].font = title_font
    
    headers = ['ID', 'Nom']
    for col, header in enumerate(headers, 1):
        cell = ws_guests.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    row = 4
    for guest in group.guest_members.all():
        ws_guests.cell(row=row, column=1, value=guest.id).border = border
        ws_guests.cell(row=row, column=2, value=guest.name).border = border
        row += 1
    
    ws_guests.column_dimensions['A'].width = 10
    ws_guests.column_dimensions['B'].width = 30
    
    # === FEUILLE 4 : DÉPENSES ===
    ws_expenses = wb.create_sheet("Depenses")
    ws_expenses['A1'] = "DÉPENSES"
    ws_expenses['A1'].font = title_font
    
    headers = ['ID', 'Titre', f'Montant ({currency_symbol})', 'Payé par', 'Date', 'Catégorie', 'Type de répartition', 'Description']
    for col, header in enumerate(headers, 1):
        cell = ws_expenses.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    row = 4
    for expense in group.expenses.select_related('paid_by').all():
        ws_expenses.cell(row=row, column=1, value=expense.id).border = border
        ws_expenses.cell(row=row, column=2, value=expense.title).border = border
        ws_expenses.cell(row=row, column=3, value=float(expense.amount)).border = border
        ws_expenses.cell(row=row, column=3).number_format = '0.00'
        ws_expenses.cell(row=row, column=4, value=expense.paid_by.username).border = border
        ws_expenses.cell(row=row, column=5, value=expense.date.strftime('%d/%m/%Y')).border = border
        ws_expenses.cell(row=row, column=6, value=expense.get_category_display()).border = border
        split_type_display = 'Égale' if expense.split_type == 'equal' else 'Personnalisée'
        ws_expenses.cell(row=row, column=7, value=split_type_display).border = border
        ws_expenses.cell(row=row, column=8, value=expense.description or '').border = border
        row += 1
    
    ws_expenses.column_dimensions['A'].width = 8
    ws_expenses.column_dimensions['B'].width = 25
    ws_expenses.column_dimensions['C'].width = 12
    ws_expenses.column_dimensions['D'].width = 20
    ws_expenses.column_dimensions['E'].width = 12
    ws_expenses.column_dimensions['F'].width = 18
    ws_expenses.column_dimensions['G'].width = 20
    ws_expenses.column_dimensions['H'].width = 35
    
    # === FEUILLE 5 : PARTICIPANTS ===
    ws_participants = wb.create_sheet("Participants")
    ws_participants['A1'] = "PARTICIPANTS PAR DÉPENSE"
    ws_participants['A1'].font = title_font
    
    headers = ['Dépense ID', 'Dépense', 'Participant', 'Type', f'Part ({currency_symbol})']
    for col, header in enumerate(headers, 1):
        cell = ws_participants.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    row = 4
    for expense in group.expenses.prefetch_related(
        'participants', 'guest_participants', 'custom_shares__user', 'custom_shares__guest'
    ).all():
        if expense.split_type == 'custom':
            for share in expense.custom_shares.all():
                participant_name = share.user.username if share.user else share.guest.name
                participant_type = 'Membre' if share.user else 'Invité'
                ws_participants.cell(row=row, column=1, value=expense.id).border = border
                ws_participants.cell(row=row, column=2, value=expense.title).border = border
                ws_participants.cell(row=row, column=3, value=participant_name).border = border
                ws_participants.cell(row=row, column=4, value=participant_type).border = border
                ws_participants.cell(row=row, column=5, value=float(share.amount)).border = border
                ws_participants.cell(row=row, column=5).number_format = '0.00'
                row += 1
        else:
            participants = list(expense.participants.all())
            guests = list(expense.guest_participants.all())
            total_count = len(participants) + len(guests)
            if total_count > 0:
                share_amount = float(expense.amount) / total_count
                for participant in participants:
                    ws_participants.cell(row=row, column=1, value=expense.id).border = border
                    ws_participants.cell(row=row, column=2, value=expense.title).border = border
                    ws_participants.cell(row=row, column=3, value=participant.username).border = border
                    ws_participants.cell(row=row, column=4, value='Membre').border = border
                    ws_participants.cell(row=row, column=5, value=share_amount).border = border
                    ws_participants.cell(row=row, column=5).number_format = '0.00'
                    row += 1
                for guest in guests:
                    ws_participants.cell(row=row, column=1, value=expense.id).border = border
                    ws_participants.cell(row=row, column=2, value=expense.title).border = border
                    ws_participants.cell(row=row, column=3, value=guest.name).border = border
                    ws_participants.cell(row=row, column=4, value='Invité').border = border
                    ws_participants.cell(row=row, column=5, value=share_amount).border = border
                    ws_participants.cell(row=row, column=5).number_format = '0.00'
                    row += 1
    
    ws_participants.column_dimensions['A'].width = 12
    ws_participants.column_dimensions['B'].width = 25
    ws_participants.column_dimensions['C'].width = 25
    ws_participants.column_dimensions['D'].width = 12
    ws_participants.column_dimensions['E'].width = 12
    
    # === FEUILLE 6 : SOLDES ===
    ws_balances = wb.create_sheet("Soldes")
    ws_balances['A1'] = "SOLDES"
    ws_balances['A1'].font = title_font
    
    headers = ['Participant', 'Type', f'Solde ({currency_symbol})']
    for col, header in enumerate(headers, 1):
        cell = ws_balances.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    row = 4
    user_balances, guest_balances = calculate_balances(group)
    for balance in user_balances:
        ws_balances.cell(row=row, column=1, value=balance['user'].username).border = border
        ws_balances.cell(row=row, column=2, value='Membre').border = border
        ws_balances.cell(row=row, column=3, value=float(balance['balance'])).border = border
        ws_balances.cell(row=row, column=3).number_format = '0.00'
        # Colorer en vert si positif, rouge si négatif
        if float(balance['balance']) > 0:
            ws_balances.cell(row=row, column=3).font = Font(color="16A34A", bold=True)
        elif float(balance['balance']) < 0:
            ws_balances.cell(row=row, column=3).font = Font(color="DC2626", bold=True)
        row += 1
    
    for balance in guest_balances:
        ws_balances.cell(row=row, column=1, value=balance['guest'].name).border = border
        ws_balances.cell(row=row, column=2, value='Invité').border = border
        ws_balances.cell(row=row, column=3, value=float(balance['balance'])).border = border
        ws_balances.cell(row=row, column=3).number_format = '0.00'
        if float(balance['balance']) > 0:
            ws_balances.cell(row=row, column=3).font = Font(color="16A34A", bold=True)
        elif float(balance['balance']) < 0:
            ws_balances.cell(row=row, column=3).font = Font(color="DC2626", bold=True)
        row += 1
    
    ws_balances.column_dimensions['A'].width = 25
    ws_balances.column_dimensions['B'].width = 12
    ws_balances.column_dimensions['C'].width = 15
    
    # === FEUILLE 7 : REMBOURSEMENTS SUGGÉRÉS ===
    ws_settlements = wb.create_sheet("Remboursements suggeres")
    ws_settlements['A1'] = "REMBOURSEMENTS SUGGÉRÉS"
    ws_settlements['A1'].font = title_font
    
    headers = ['De', 'Type débiteur', 'À', f'Montant ({currency_symbol})']
    for col, header in enumerate(headers, 1):
        cell = ws_settlements.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    row = 4
    settlements = calculate_settlements(group)
    for settlement in settlements:
        from_name = settlement['from']['obj'].username if settlement['from']['type'] == 'user' else settlement['from']['obj'].name
        to_name = settlement['to']['obj'].username
        from_type = 'Membre' if settlement['from']['type'] == 'user' else 'Invité'
        
        ws_settlements.cell(row=row, column=1, value=from_name).border = border
        ws_settlements.cell(row=row, column=2, value=from_type).border = border
        ws_settlements.cell(row=row, column=3, value=to_name).border = border
        ws_settlements.cell(row=row, column=4, value=float(settlement['amount'])).border = border
        ws_settlements.cell(row=row, column=4).number_format = '0.00'
        ws_settlements.cell(row=row, column=4).font = Font(color="4F46E5", bold=True)
        row += 1
    
    ws_settlements.column_dimensions['A'].width = 25
    ws_settlements.column_dimensions['B'].width = 15
    ws_settlements.column_dimensions['C'].width = 25
    ws_settlements.column_dimensions['D'].width = 15
    
    # === FEUILLE 8 : REMBOURSEMENTS EFFECTUÉS ===
    ws_reimbursements = wb.create_sheet("Remboursements effectues")
    ws_reimbursements['A1'] = "REMBOURSEMENTS EFFECTUÉS"
    ws_reimbursements['A1'].font = title_font
    
    headers = ['De', 'Type', 'À', f'Montant ({currency_symbol})', 'Date']
    for col, header in enumerate(headers, 1):
        cell = ws_reimbursements.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    row = 4
    for reimbursement in group.reimbursements.select_related('from_user', 'from_guest', 'to_user').all():
        from_name = reimbursement.from_user.username if reimbursement.from_user else reimbursement.from_guest.name
        from_type = 'Membre' if reimbursement.from_user else 'Invité'
        
        ws_reimbursements.cell(row=row, column=1, value=from_name).border = border
        ws_reimbursements.cell(row=row, column=2, value=from_type).border = border
        ws_reimbursements.cell(row=row, column=3, value=reimbursement.to_user.username).border = border
        ws_reimbursements.cell(row=row, column=4, value=float(reimbursement.amount)).border = border
        ws_reimbursements.cell(row=row, column=4).number_format = '0.00'
        ws_reimbursements.cell(row=row, column=5, value=reimbursement.created_at.strftime('%d/%m/%Y %H:%M')).border = border
        row += 1
    
    ws_reimbursements.column_dimensions['A'].width = 25
    ws_reimbursements.column_dimensions['B'].width = 12
    ws_reimbursements.column_dimensions['C'].width = 25
    ws_reimbursements.column_dimensions['D'].width = 15
    ws_reimbursements.column_dimensions['E'].width = 18
    
    # Sauvegarder dans un buffer
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output
